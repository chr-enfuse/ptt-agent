"""Workbook upload + structure endpoints and the in-process workbook store —
the Python port of backend/src/routes/workbooks.ts.

openpyxl needs two loads of each upload: `data_only=True` to read cached formula
results (matches exceljs `formula.result`), and `data_only=False` to read the
formula strings themselves. We keep both, keyed by file_id. Numbers are always
re-read from here by deterministic code — they never round-trip through model text.
"""

from __future__ import annotations

import io
import uuid
from dataclasses import dataclass

import openpyxl
from fastapi import APIRouter, File, UploadFile
from fastapi.responses import JSONResponse

from .excel import describe_structure

router = APIRouter()

# 25 MB upload cap, matching the Node multer limit.
MAX_UPLOAD_BYTES = 25 * 1024 * 1024


@dataclass
class LoadedWorkbook:
    """A parsed upload: `values` carries cached formula results (data_only), while
    `formulas` carries the formula strings."""

    values: openpyxl.Workbook
    formulas: openpyxl.Workbook


# In-memory store for v1 scaffolding (mirrors the Node Map<string, Workbook>).
_workbooks: dict[str, LoadedWorkbook] = {}


def get_workbook(file_id: str) -> LoadedWorkbook | None:
    return _workbooks.get(file_id)


@router.post("")
@router.post("/")
async def upload_workbook(file: UploadFile | None = File(default=None)) -> JSONResponse:
    """POST /api/workbooks (multipart/form-data, field name: "file").
    Parses the uploaded .xlsx deterministically and returns { fileId, filename, sheets }."""
    if file is None:
        return JSONResponse(status_code=400, content={"error": "multipart field `file` is required (.xlsx)"})

    data = await file.read()
    if len(data) > MAX_UPLOAD_BYTES:
        return JSONResponse(status_code=400, content={"error": "file exceeds the 25 MB limit"})

    try:
        values = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
        formulas = openpyxl.load_workbook(io.BytesIO(data), data_only=False)
    except Exception as err:  # openpyxl raises a variety of parse errors
        return JSONResponse(status_code=422, content={"error": f"failed to parse workbook: {err}"})

    file_id = str(uuid.uuid4())
    _workbooks[file_id] = LoadedWorkbook(values=values, formulas=formulas)

    return JSONResponse(
        content={
            "fileId": file_id,
            "filename": file.filename,
            "sheets": [
                {"name": ws.title, "rowCount": ws.max_row or 0, "columnCount": ws.max_column or 0}
                for ws in values.worksheets
            ],
        }
    )


@router.get("/{file_id}/structure")
async def workbook_structure(file_id: str) -> JSONResponse:
    """GET /api/workbooks/:fileId/structure — sheets + headers (no bulk values)."""
    wb = get_workbook(file_id)
    if wb is None:
        return JSONResponse(status_code=404, content={"error": "unknown fileId"})
    return JSONResponse(content=describe_structure(wb))
