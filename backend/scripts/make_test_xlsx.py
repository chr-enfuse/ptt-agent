"""Create a test .xlsx with varied number formats for Node/Python display-text
parity checks (plan §7.3, acceptance case 4)."""

import sys

import openpyxl

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Financials"

rows = [
    ("Revenue", 13456250, "#,##0"),
    ("Revenue $", 13456250, '"$"#,##0'),
    ("Margin", 0.153, "0.0%"),
    ("Margin pp", 0.153, "0%"),
    ("Cost", 1234500.5, "#,##0.00"),
    ("Cost $", 1234500.5, '"$"#,##0.00'),
    ("Units", 4200, "0"),
    ("Ratio", 1.875, "0.00"),
    ("Loss", -4200, "#,##0;(#,##0)"),
    ("Label", "Acme Corp", "General"),
    ("Plain", 1234500, "General"),
]

ws["A1"] = "Metric"
ws["B1"] = "Value"
for i, (label, value, fmt) in enumerate(rows, start=2):
    ws.cell(row=i, column=1, value=label)
    c = ws.cell(row=i, column=2, value=value)
    c.number_format = fmt

out = sys.argv[1] if len(sys.argv) > 1 else "test.xlsx"
wb.save(out)
print(f"wrote {out} with {len(rows)} value rows (B2:B{len(rows) + 1})")
